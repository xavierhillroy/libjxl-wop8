# src/reporting/spreadsheet.py
import os
import sys
import pandas as pd

# Add parent directory to path to find config
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..')))
from config import SPREADSHEETS_DIR

def create_dataset_spreadsheet(train_stats, test_stats, run_name):
    """
    Create a spreadsheet with multiple sheets for a dataset.
    
    Args:
        train_stats (list): List of dictionaries with training image statistics
        test_stats (list): List of dictionaries with testing image statistics
        dataset_name (str): Name of the dataset
    
    Returns:
        str: Path to the created spreadsheet
    """
    # Create directory if it doesn't exist
    os.makedirs(SPREADSHEETS_DIR, exist_ok=True)
    excel_path = os.path.join(SPREADSHEETS_DIR, f"{run_name}_results.xlsx")
    
    # Define columns for training sheet
    train_columns = [
        'image_name',
        'width',
        'height',
        'num_pixels',
        'uncompressed_size_bytes',
        'baseline_size_bytes',
        'baseline_bpp',
        'baseline_mae'
    ]
    
    # Define columns for testing sheet
    test_columns = [
        'image_name',
        'width',
        'height',
        'num_pixels',
        'uncompressed_size_bytes',
        'baseline_size_bytes',
        'baseline_bpp',
        'baseline_mae',
        'wop8_size_bytes',
        'wop8_bpp',
        'wop8_mae',
        'size_reduction_bytes',
        'bpp_improvement',
        'improvement_percentage'
    ]
    
    # Create training DataFrame and ensure all columns exist
    train_df = pd.DataFrame(train_stats)
    for col in train_columns:
        if col not in train_df.columns:
            train_df[col] = None
    
    # Ensure columns are in the right order
    train_df = train_df[train_columns]
    
    # Create testing DataFrame and ensure all columns exist
    test_df = pd.DataFrame(test_stats)
    for col in test_columns:
        if col not in test_df.columns:
            test_df[col] = None
    
    # Ensure columns are in the right order
    test_df = test_df[test_columns]
    
    # Create all images DataFrame with the same structure as testing
    all_stats = train_stats + test_stats
    all_df = pd.DataFrame(all_stats)
    for col in test_columns:
        if col not in all_df.columns:
            all_df[col] = None
    
    # Ensure columns are in the right order
    all_df = all_df[test_columns]
    
    # Create effort level sheets with the same structure as All Images
    effort7_df = pd.DataFrame(all_stats)
    effort9_df = pd.DataFrame(all_stats)
    
    for col in test_columns:
        if col not in effort7_df.columns:
            effort7_df[col] = None
            effort9_df[col] = None
    
    # Ensure columns are in the right order
    effort7_df = effort7_df[test_columns]
    effort9_df = effort9_df[test_columns]
    
    # Write to Excel file with multiple sheets
    with pd.ExcelWriter(excel_path) as writer:
        train_df.to_excel(writer, sheet_name='Training', index=False)
        test_df.to_excel(writer, sheet_name='Testing', index=False)
        all_df.to_excel(writer, sheet_name='All Images', index=False)
        effort7_df.to_excel(writer, sheet_name='Effort Level 7', index=False)
        effort9_df.to_excel(writer, sheet_name='Effort Level 9', index=False)
        
        # Add a totals row to each sheet
        for sheet_name in ['Training', 'Testing', 'All Images', 'Effort Level 7', 'Effort Level 9']:
            worksheet = writer.sheets[sheet_name]
            num_rows = len(train_df) if sheet_name == 'Training' else (len(test_df) if sheet_name == 'Testing' else len(all_df))
            worksheet.cell(row=num_rows+2, column=1, value="TOTAL")
            
            # Add SUM formulas for baseline size columns (bytes) and pixel columns
            size_columns = ['E']  # uncompressed_size_bytes, baseline_size_bytes
            pixel_columns = ['D']  # num_pixels
            
            # Calculate sums for both size and pixel columns
            for col in pixel_columns + size_columns:
                column_index = ord(col) - ord('A') + 1
                columns_in_sheet = len(train_columns) if sheet_name == 'Training' else len(test_columns)
                
                if column_index <= columns_in_sheet:
                    formula = f"=SUM({col}2:{col}{num_rows+1})"
                    worksheet.cell(row=num_rows+2, column=column_index, value=formula)
            
            # Note: BPP calculations will be added later when baseline data is available
            # through update_spreadsheet_with_baseline function
            
            # Note: MAE calculations will be added later when baseline data is available
            # through update_spreadsheet_with_baseline function
    
    return excel_path

def update_spreadsheet_with_ga_candidate(excel_path, candidate, results):
    """
    Update the training spreadsheet with a single GA candidate's results.
    Ensures proper total row handling.
    
    Args:
        excel_path (str): Path to the Excel file
        candidate (list): List of weights [w0, w1, w2, ...]
        results (dict): Dictionary with results {image_name: {'size': size, 'mae': mae}}
        
    Returns:
        bool: True if successful, False otherwise
    """
    try:
        # Read existing Excel file
        train_df = pd.read_excel(excel_path, sheet_name='Training')
        
        # Store total row before updating
        has_total = 'TOTAL' in train_df['image_name'].values
        if has_total:
            total_row = train_df[train_df['image_name'] == 'TOTAL'].copy()
            train_df = train_df[train_df['image_name'] != 'TOTAL']
        
        # Create column names for this candidate
        weight_str = '_'.join(map(str, candidate))
        column_prefix = f"w{weight_str}"
        
        # Check if these columns already exist
        if f"{column_prefix}_fitness" in train_df.columns:
            print(f"Candidate {weight_str} already in spreadsheet, skipping")
            return True  # Already in spreadsheet
        
        # Add columns for this weight configuration (only fitness and mae as requested)
        train_df[f"{column_prefix}_fitness"] = None
        train_df[f"{column_prefix}_mae"] = None
        
        # Fill data for each image
        for img_name, result in results.items():
            size = result['size']
            mae = result['mae']
            
            # Update row for this image
            mask = train_df['image_name'] == img_name
            if any(mask):
                train_df.loc[mask, f"{column_prefix}_fitness"] = -size
                train_df.loc[mask, f"{column_prefix}_mae"] = mae
        
        # Calculate and add totals back
        if has_total:
            # Add the new columns to the total row
            total_row[f"{column_prefix}_fitness"] = train_df[f"{column_prefix}_fitness"].sum()
            total_row[f"{column_prefix}_mae"] = train_df[f"{column_prefix}_mae"].mean()
            
            # Add the total row back to the dataframe
            train_df = pd.concat([train_df, total_row], ignore_index=True)
        else:
            # Create a new total row with all columns
            total_data = {'image_name': 'TOTAL'}
            for col in train_df.columns:
                if col == 'image_name':
                    continue
                elif col.endswith('_fitness'):
                    total_data[col] = train_df[col].sum()
                elif col.endswith('_mae'):
                    total_data[col] = train_df[col].mean()
                elif col in ['num_pixels', 'uncompressed_size_bytes']:
                    total_data[col] = train_df[col].sum()
                elif col in ['baseline_size_bytes'] and col in train_df.columns:
                    total_data[col] = train_df[col].sum()
                elif col in ['baseline_bpp'] and col in train_df.columns:
                    total_data[col] = train_df[col].mean()
                else:
                    total_data[col] = None
                    
            total_row = pd.DataFrame([total_data])
            train_df = pd.concat([train_df, total_row], ignore_index=True)
        
        # Write updated dataframe back to Excel
        with pd.ExcelWriter(excel_path, mode='a', if_sheet_exists='replace') as writer:
            train_df.to_excel(writer, sheet_name='Training', index=False)
        
        return True
    except Exception as e:
        print(f"Error updating spreadsheet with GA candidate: {e}")
        return False

def update_with_wop8_results(excel_path, wop8_results):
    """
    Update testing and all images sheets with W-OP8 results.
    
    Args:
        excel_path (str): Path to the Excel file
        wop8_results (dict): Dictionary with W-OP8 results
            {
                'results': [
                    {
                        'image_name': 'img1.png', 
                        'size': 1000, 
                        'mae': 0.0
                    },
                    ...
                ]
            }
    
    Returns:
        bool: True if successful, False otherwise
    """
    try:
        # Read existing Excel file
        test_df = pd.read_excel(excel_path, sheet_name='Testing')
        all_df = pd.read_excel(excel_path, sheet_name='All Images')
        
        # Store total rows before updating
        test_total = test_df[test_df['image_name'] == 'TOTAL'].copy() if 'TOTAL' in test_df['image_name'].values else None
        all_total = all_df[all_df['image_name'] == 'TOTAL'].copy() if 'TOTAL' in all_df['image_name'].values else None
        
        # Remove total rows for updating
        test_df = test_df[test_df['image_name'] != 'TOTAL']
        all_df = all_df[all_df['image_name'] != 'TOTAL']
        
        # Update dataframes with W-OP8 results
        for result in wop8_results['results']:
            img_name = result['image_name']
            size = result['size']
            mae = result['mae']
            
            # Update testing sheet
            test_mask = test_df['image_name'] == img_name
            if any(test_mask):
                test_df.loc[test_mask, 'wop8_size_bytes'] = size
                test_df.loc[test_mask, 'wop8_mae'] = mae
                
                # Calculate derived metrics
                baseline_size = test_df.loc[test_mask, 'baseline_size_bytes'].values[0]
                num_pixels = test_df.loc[test_mask, 'num_pixels'].values[0]
                
                # Calculate W-OP8 bits per pixel correctly
                wop8_bpp = (size * 8) / num_pixels
                test_df.loc[test_mask, 'wop8_bpp'] = wop8_bpp
                
                # Calculate difference vs baseline
                size_reduction = baseline_size - size
                test_df.loc[test_mask, 'size_reduction_bytes'] = size_reduction
                
                # Calculate bpp improvement
                baseline_bpp = test_df.loc[test_mask, 'baseline_bpp'].values[0]
                bpp_improvement = baseline_bpp - wop8_bpp
                test_df.loc[test_mask, 'bpp_improvement'] = bpp_improvement
                
                # Calculate improvement percentage
                improvement_percentage = (size_reduction / baseline_size) * 100
                test_df.loc[test_mask, 'improvement_percentage'] = improvement_percentage
            
            # Update all images sheet
            all_mask = all_df['image_name'] == img_name
            if any(all_mask):
                all_df.loc[all_mask, 'wop8_size_bytes'] = size
                all_df.loc[all_mask, 'wop8_mae'] = mae
                
                # Calculate derived metrics
                baseline_size = all_df.loc[all_mask, 'baseline_size_bytes'].values[0]
                num_pixels = all_df.loc[all_mask, 'num_pixels'].values[0]
                
                # Calculate W-OP8 bits per pixel correctly
                wop8_bpp = (size * 8) / num_pixels
                all_df.loc[all_mask, 'wop8_bpp'] = wop8_bpp
                
                # Calculate difference vs baseline
                size_reduction = baseline_size - size
                all_df.loc[all_mask, 'size_reduction_bytes'] = size_reduction
                
                # Calculate bpp improvement
                baseline_bpp = all_df.loc[all_mask, 'baseline_bpp'].values[0]
                bpp_improvement = baseline_bpp - wop8_bpp
                all_df.loc[all_mask, 'bpp_improvement'] = bpp_improvement
                
                # Calculate improvement percentage
                improvement_percentage = (size_reduction / baseline_size) * 100
                all_df.loc[all_mask, 'improvement_percentage'] = improvement_percentage
        
        # Recalculate totals for testing sheet with correct BPP calculation
        test_total_row = pd.DataFrame({
            'image_name': ['TOTAL'],
            'width': [None],
            'height': [None],
            'num_pixels': [test_df['num_pixels'].sum()],
            'uncompressed_size_bytes': [test_df['uncompressed_size_bytes'].sum()],
            'baseline_size_bytes': [test_df['baseline_size_bytes'].sum()],
            'baseline_mae': [test_df['baseline_mae'].mean()],
            'wop8_size_bytes': [test_df['wop8_size_bytes'].sum()],
            'wop8_mae': [test_df['wop8_mae'].mean()],
            'size_reduction_bytes': [test_df['size_reduction_bytes'].sum()],
        })
        
        # Calculate BPP based on total dataset values
        total_pixels = test_total_row['num_pixels'].values[0]
        baseline_total_size = test_total_row['baseline_size_bytes'].values[0]
        wop8_total_size = test_total_row['wop8_size_bytes'].values[0]
        
        # Calculate BPP for total dataset
        baseline_bpp = (baseline_total_size * 8) / total_pixels
        wop8_bpp = (wop8_total_size * 8) / total_pixels
        
        # Calculate BPP improvement and percentage based on totals
        bpp_improvement = baseline_bpp - wop8_bpp
        improvement_percentage = (test_total_row['size_reduction_bytes'].values[0] / baseline_total_size) * 100
        
        # Add these calculated values to the total row
        test_total_row['baseline_bpp'] = baseline_bpp
        test_total_row['wop8_bpp'] = wop8_bpp
        test_total_row['bpp_improvement'] = bpp_improvement
        test_total_row['improvement_percentage'] = improvement_percentage
        
        # Similar calculations for all images sheet
        all_total_row = pd.DataFrame({
            'image_name': ['TOTAL'],
            'width': [None],
            'height': [None],
            'num_pixels': [all_df['num_pixels'].sum()],
            'uncompressed_size_bytes': [all_df['uncompressed_size_bytes'].sum()],
            'baseline_size_bytes': [all_df['baseline_size_bytes'].sum()],
            'baseline_mae': [all_df['baseline_mae'].mean()],
            'wop8_size_bytes': [all_df['wop8_size_bytes'].sum()],
            'wop8_mae': [all_df['wop8_mae'].mean()],
            'size_reduction_bytes': [all_df['size_reduction_bytes'].sum()],
        })
        
        # Calculate BPP based on total dataset values for all images
        total_pixels = all_total_row['num_pixels'].values[0]
        baseline_total_size = all_total_row['baseline_size_bytes'].values[0]
        wop8_total_size = all_total_row['wop8_size_bytes'].values[0]
        
        # Calculate BPP for total dataset
        baseline_bpp = (baseline_total_size * 8) / total_pixels
        wop8_bpp = (wop8_total_size * 8) / total_pixels
        
        # Calculate BPP improvement and percentage based on totals
        bpp_improvement = baseline_bpp - wop8_bpp
        improvement_percentage = (all_total_row['size_reduction_bytes'].values[0] / baseline_total_size) * 100
        
        # Add these calculated values to the total row
        all_total_row['baseline_bpp'] = baseline_bpp
        all_total_row['wop8_bpp'] = wop8_bpp
        all_total_row['bpp_improvement'] = bpp_improvement
        all_total_row['improvement_percentage'] = improvement_percentage
        
        # Append total rows back to dataframes
        test_df = pd.concat([test_df, test_total_row], ignore_index=True)
        all_df = pd.concat([all_df, all_total_row], ignore_index=True)
        
        # Write updated dataframes back to Excel
        with pd.ExcelWriter(excel_path, mode='a', if_sheet_exists='replace') as writer:
            test_df.to_excel(writer, sheet_name='Testing', index=False)
            all_df.to_excel(writer, sheet_name='All Images', index=False)
        
        return True
    except Exception as e:
        print(f"Error updating with W-OP8 results: {e}")
        return False

def update_spreadsheet_with_baseline(excel_path, train_results, test_results):
    """
    Update the spreadsheet with baseline compression results.
    
    Args:
        excel_path (str): Path to the Excel file
        train_results (dict): Dictionary with training results {image_name: {size, mae}}
        test_results (dict): Dictionary with testing results {image_name: {size, mae}}
        
    Returns:
        bool: True if successful, False otherwise
    """
    try:
        # Read existing Excel file
        train_df = pd.read_excel(excel_path, sheet_name='Training')
        test_df = pd.read_excel(excel_path, sheet_name='Testing')
        all_df = pd.read_excel(excel_path, sheet_name='All Images')
        
        # Store total rows before updating
        train_total = train_df[train_df['image_name'] == 'TOTAL'].copy() if 'TOTAL' in train_df['image_name'].values else None
        test_total = test_df[test_df['image_name'] == 'TOTAL'].copy() if 'TOTAL' in test_df['image_name'].values else None
        all_total = all_df[all_df['image_name'] == 'TOTAL'].copy() if 'TOTAL' in all_df['image_name'].values else None
        
        # Remove total rows for updating
        train_df = train_df[train_df['image_name'] != 'TOTAL']
        test_df = test_df[test_df['image_name'] != 'TOTAL']
        all_df = all_df[all_df['image_name'] != 'TOTAL']
        
        # Update training sheet
        for img_name, result in train_results.items():
            mask = train_df['image_name'] == img_name
            if any(mask):
                compressed_size = result['size']
                mae = result['mae']
                
                train_df.loc[mask, 'baseline_size_bytes'] = compressed_size
                train_df.loc[mask, 'baseline_mae'] = mae
                
                # Calculate bits per pixel
                num_pixels = train_df.loc[mask, 'num_pixels'].values[0]
                bpp = (compressed_size * 8) / num_pixels
                train_df.loc[mask, 'baseline_bpp'] = bpp
                
                # Update all_df as well
                all_mask = all_df['image_name'] == img_name
                if any(all_mask):
                    all_df.loc[all_mask, 'baseline_size_bytes'] = compressed_size
                    all_df.loc[all_mask, 'baseline_mae'] = mae
                    all_df.loc[all_mask, 'baseline_bpp'] = bpp
        
        # Update testing sheet
        for img_name, result in test_results.items():
            mask = test_df['image_name'] == img_name
            if any(mask):
                compressed_size = result['size']
                mae = result['mae']
                
                test_df.loc[mask, 'baseline_size_bytes'] = compressed_size
                test_df.loc[mask, 'baseline_mae'] = mae
                
                # Calculate bits per pixel
                num_pixels = test_df.loc[mask, 'num_pixels'].values[0]
                bpp = (compressed_size * 8) / num_pixels
                test_df.loc[mask, 'baseline_bpp'] = bpp
                
                # Update all_df as well
                all_mask = all_df['image_name'] == img_name
                if any(all_mask):
                    all_df.loc[all_mask, 'baseline_size_bytes'] = compressed_size
                    all_df.loc[all_mask, 'baseline_mae'] = mae
                    all_df.loc[all_mask, 'baseline_bpp'] = bpp
        
        # Calculate and add totals back
        # Training totals
        if train_total is not None:
            # Calculate new totals
            total_pixels = train_df['num_pixels'].sum()
            total_baseline_size = train_df['baseline_size_bytes'].sum()
            baseline_bpp_total = (total_baseline_size * 8) / total_pixels
            
            train_total_row = pd.DataFrame({
                'image_name': ['TOTAL'],
                'width': [None],
                'height': [None],
                'num_pixels': [total_pixels],
                'uncompressed_size_bytes': [train_df['uncompressed_size_bytes'].sum()],
                'baseline_size_bytes': [total_baseline_size],
                'baseline_bpp': [baseline_bpp_total],
                'baseline_mae': [train_df['baseline_mae'].mean()]
            })
            # Add other columns that may exist
            for col in train_df.columns:
                if col not in train_total_row.columns:
                    train_total_row[col] = None
            train_df = pd.concat([train_df, train_total_row], ignore_index=True)
        
        # Testing totals
        if test_total is not None:
            # Calculate new totals
            total_pixels = test_df['num_pixels'].sum()
            total_baseline_size = test_df['baseline_size_bytes'].sum()
            baseline_bpp_total = (total_baseline_size * 8) / total_pixels
            
            test_total_row = pd.DataFrame({
                'image_name': ['TOTAL'],
                'width': [None],
                'height': [None],
                'num_pixels': [total_pixels],
                'uncompressed_size_bytes': [test_df['uncompressed_size_bytes'].sum()],
                'baseline_size_bytes': [total_baseline_size],
                'baseline_bpp': [baseline_bpp_total],
                'baseline_mae': [test_df['baseline_mae'].mean()]
            })
            # Add other columns that may exist
            for col in test_df.columns:
                if col not in test_total_row.columns:
                    test_total_row[col] = None
            test_df = pd.concat([test_df, test_total_row], ignore_index=True)
        
        # All images totals
        if all_total is not None:
            # Calculate new totals
            total_pixels = all_df['num_pixels'].sum()
            total_baseline_size = all_df['baseline_size_bytes'].sum()
            baseline_bpp_total = (total_baseline_size * 8) / total_pixels
            
            all_total_row = pd.DataFrame({
                'image_name': ['TOTAL'],
                'width': [None],
                'height': [None],
                'num_pixels': [total_pixels],
                'uncompressed_size_bytes': [all_df['uncompressed_size_bytes'].sum()],
                'baseline_size_bytes': [total_baseline_size],
                'baseline_bpp': [baseline_bpp_total],
                'baseline_mae': [all_df['baseline_mae'].mean()]
            })
            # Add other columns that may exist
            for col in all_df.columns:
                if col not in all_total_row.columns:
                    all_total_row[col] = None
            all_df = pd.concat([all_df, all_total_row], ignore_index=True)
        
        # Write updated dataframes back to Excel
        with pd.ExcelWriter(excel_path, mode='a', if_sheet_exists='replace') as writer:
            train_df.to_excel(writer, sheet_name='Training', index=False)
            test_df.to_excel(writer, sheet_name='Testing', index=False)
            all_df.to_excel(writer, sheet_name='All Images', index=False)
        
        return True
    except Exception as e:
        print(f"Error updating spreadsheet with baseline results: {e}")
        return False

def create_summary_sheet(excel_path):
    """
    Create a summary sheet with key metrics comparing baseline and W-OP8 performance.
    
    Args:
        excel_path (str): Path to the Excel file
        
    Returns:
        bool: True if successful, False otherwise
    """
    try:
        # Read existing Excel file
        test_df = pd.read_excel(excel_path, sheet_name='Testing')
        all_df = pd.read_excel(excel_path, sheet_name='All Images')
        
        # Get total rows
        test_total = test_df[test_df['image_name'] == 'TOTAL'].iloc[0]
        all_total = all_df[all_df['image_name'] == 'TOTAL'].iloc[0]
        
        # Extract metrics from testing set totals
        test_metrics = {
            'baseline_total_size': test_total['baseline_size_bytes'],
            'baseline_bpp': test_total['baseline_bpp'],
        }
        
        # Add W-OP8 metrics if they exist
        if 'wop8_size_bytes' in test_total:
            test_metrics.update({
                'wop8_total_size': test_total['wop8_size_bytes'],
                'wop8_bpp': test_total['wop8_bpp'],
                'size_reduction': test_total['size_reduction_bytes'],
                'improvement_percentage': test_total['improvement_percentage'],
                'bpp_improvement': test_total['bpp_improvement']
            })
        else:
            # Set default values if W-OP8 data doesn't exist
            test_metrics.update({
                'wop8_total_size': 'N/A',
                'wop8_bpp': 'N/A',
                'size_reduction': 'N/A',
                'improvement_percentage': 'N/A',
                'bpp_improvement': 'N/A'
            })
        
        # Extract metrics from all images totals
        all_metrics = {
            'baseline_total_size': all_total['baseline_size_bytes'],
            'baseline_bpp': all_total['baseline_bpp'],
        }
        
        # Add W-OP8 metrics if they exist
        if 'wop8_size_bytes' in all_total:
            all_metrics.update({
                'wop8_total_size': all_total['wop8_size_bytes'],
                'wop8_bpp': all_total['wop8_bpp'],
                'size_reduction': all_total['size_reduction_bytes'],
                'improvement_percentage': all_total['improvement_percentage'],
                'bpp_improvement': all_total['bpp_improvement']
            })
        else:
            # Set default values if W-OP8 data doesn't exist
            all_metrics.update({
                'wop8_total_size': 'N/A',
                'wop8_bpp': 'N/A',
                'size_reduction': 'N/A',
                'improvement_percentage': 'N/A',
                'bpp_improvement': 'N/A'
            })
        
        # Create summary dataframe
        summary_data = [
            ['Metric', 'Testing Set', 'All Images'],
            ['', '', ''],
            ['Dataset Statistics:', '', ''],
            ['Images Count', len(test_df) - 1, len(all_df) - 1],  # Subtract 1 for total row
            ['', '', ''],
            ['Baseline Performance:', '', ''],
            ['Total Size (bytes)', f"{test_metrics['baseline_total_size']:,}", f"{all_metrics['baseline_total_size']:,}"],
            ['Bits per Pixel', f"{test_metrics['baseline_bpp']:.3f}", f"{all_metrics['baseline_bpp']:.3f}"],
            ['', '', ''],
            ['W-OP8 Performance:', '', ''],
            ['Total Size (bytes)', str(test_metrics['wop8_total_size']), str(all_metrics['wop8_total_size'])],
            ['Bits per Pixel', str(test_metrics['wop8_bpp']), str(all_metrics['wop8_bpp'])],
            ['', '', ''],
            ['Improvements:', '', ''],
            ['Size Reduction (bytes)', str(test_metrics['size_reduction']), str(all_metrics['size_reduction'])],
            ['Size Reduction (%)', str(test_metrics['improvement_percentage']), str(all_metrics['improvement_percentage'])],
            ['Bits per Pixel Improvement', str(test_metrics['bpp_improvement']), str(all_metrics['bpp_improvement'])],
        ]
        
        # Convert to DataFrame
        summary_df = pd.DataFrame(summary_data)
        summary_df.columns = summary_data[0]
        summary_df = summary_df[1:]  # Remove header row
        
        # Read the existing workbook to preserve all sheets
        with pd.ExcelWriter(excel_path, mode='a', if_sheet_exists='replace') as writer:
            # Write summary sheet
            summary_df.to_excel(writer, sheet_name='Summary', index=False, header=False)
            
            # Access the workbook and set column widths for better formatting
            worksheet = writer.sheets['Summary']
            worksheet.column_dimensions['A'].width = 30
            worksheet.column_dimensions['B'].width = 20
            worksheet.column_dimensions['C'].width = 20
            
            # Add some basic styling to make it more readable
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            
            # Style header row
            for col in range(1, 4):  # Columns A, B, C
                cell = worksheet.cell(row=1, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # Style section headers
            section_rows = [3, 6, 10, 14]  # Rows with section headers
            for row in section_rows:
                worksheet.cell(row=row, column=1).font = Font(bold=True)
        
        return True
    except Exception as e:
        print(f"Error creating summary sheet: {e}")
        return False

def update_with_effort_results(excel_path, effort_results):
    """
    Update effort level sheets with compression results.
    
    Args:
        excel_path (str): Path to the Excel file
        effort_results (dict): Dictionary with effort level results
            {
                'effort7': {
                    'baseline': {image_name: {'size': size, 'mae': mae}, ...},
                    'wop8': {image_name: {'size': size, 'mae': mae}, ...}
                },
                'effort9': {
                    'baseline': {image_name: {'size': size, 'mae': mae}, ...},
                    'wop8': {image_name: {'size': size, 'mae': mae}, ...}
                }
            }
    
    Returns:
        bool: True if successful, False otherwise
    """
    try:
        # Read existing Excel file
        effort7_df = pd.read_excel(excel_path, sheet_name='Effort Level 7')
        effort9_df = pd.read_excel(excel_path, sheet_name='Effort Level 9')
        
        # Store total rows before updating
        effort7_total = effort7_df[effort7_df['image_name'] == 'TOTAL'].copy() if 'TOTAL' in effort7_df['image_name'].values else None
        effort9_total = effort9_df[effort9_df['image_name'] == 'TOTAL'].copy() if 'TOTAL' in effort9_df['image_name'].values else None
        
        # Remove total rows for updating
        effort7_df = effort7_df[effort7_df['image_name'] != 'TOTAL']
        effort9_df = effort9_df[effort9_df['image_name'] != 'TOTAL']
        
        # Update Effort Level 7 sheet
        baseline_effort7 = effort_results['effort7']['baseline']
        wop8_effort7 = effort_results['effort7']['wop8']
        
        for img_name, result in baseline_effort7.items():
            mask = effort7_df['image_name'] == img_name
            if any(mask):
                # Update baseline values
                effort7_df.loc[mask, 'baseline_size_bytes'] = result['size']
                effort7_df.loc[mask, 'baseline_mae'] = result['mae']
                
                # Calculate bits per pixel
                num_pixels = effort7_df.loc[mask, 'num_pixels'].values[0]
                baseline_bpp = (result['size'] * 8) / num_pixels
                effort7_df.loc[mask, 'baseline_bpp'] = baseline_bpp
                
                # Update W-OP8 values if available
                if img_name in wop8_effort7:
                    wop8_result = wop8_effort7[img_name]
                    effort7_df.loc[mask, 'wop8_size_bytes'] = wop8_result['size']
                    effort7_df.loc[mask, 'wop8_mae'] = wop8_result['mae']
                    
                    # Calculate derived metrics
                    wop8_bpp = (wop8_result['size'] * 8) / num_pixels
                    effort7_df.loc[mask, 'wop8_bpp'] = wop8_bpp
                    
                    # Calculate differences
                    size_reduction = result['size'] - wop8_result['size']
                    effort7_df.loc[mask, 'size_reduction_bytes'] = size_reduction
                    
                    bpp_improvement = baseline_bpp - wop8_bpp
                    effort7_df.loc[mask, 'bpp_improvement'] = bpp_improvement
                    
                    improvement_percentage = (size_reduction / result['size']) * 100
                    effort7_df.loc[mask, 'improvement_percentage'] = improvement_percentage
        
        # Update Effort Level 9 sheet
        baseline_effort9 = effort_results['effort9']['baseline']
        wop8_effort9 = effort_results['effort9']['wop8']
        
        for img_name, result in baseline_effort9.items():
            mask = effort9_df['image_name'] == img_name
            if any(mask):
                # Update baseline values
                effort9_df.loc[mask, 'baseline_size_bytes'] = result['size']
                effort9_df.loc[mask, 'baseline_mae'] = result['mae']
                
                # Calculate bits per pixel
                num_pixels = effort9_df.loc[mask, 'num_pixels'].values[0]
                baseline_bpp = (result['size'] * 8) / num_pixels
                effort9_df.loc[mask, 'baseline_bpp'] = baseline_bpp
                
                # Update W-OP8 values if available
                if img_name in wop8_effort9:
                    wop8_result = wop8_effort9[img_name]
                    effort9_df.loc[mask, 'wop8_size_bytes'] = wop8_result['size']
                    effort9_df.loc[mask, 'wop8_mae'] = wop8_result['mae']
                    
                    # Calculate derived metrics
                    wop8_bpp = (wop8_result['size'] * 8) / num_pixels
                    effort9_df.loc[mask, 'wop8_bpp'] = wop8_bpp
                    
                    # Calculate differences
                    size_reduction = result['size'] - wop8_result['size']
                    effort9_df.loc[mask, 'size_reduction_bytes'] = size_reduction
                    
                    bpp_improvement = baseline_bpp - wop8_bpp
                    effort9_df.loc[mask, 'bpp_improvement'] = bpp_improvement
                    
                    improvement_percentage = (size_reduction / result['size']) * 100
                    effort9_df.loc[mask, 'improvement_percentage'] = improvement_percentage
        
        # Calculate totals for Effort Level 7 with correct BPP calculation
        effort7_total_row = pd.DataFrame({
            'image_name': ['TOTAL'],
            'width': [None],
            'height': [None],
            'num_pixels': [effort7_df['num_pixels'].sum()],
            'uncompressed_size_bytes': [effort7_df['uncompressed_size_bytes'].sum()],
            'baseline_size_bytes': [effort7_df['baseline_size_bytes'].sum()],
            'baseline_mae': [effort7_df['baseline_mae'].mean()],
            'wop8_size_bytes': [effort7_df['wop8_size_bytes'].sum()],
            'wop8_mae': [effort7_df['wop8_mae'].mean()],
            'size_reduction_bytes': [effort7_df['size_reduction_bytes'].sum()],
        })
        
        # Calculate BPP based on total dataset values for effort 7
        total_pixels = effort7_total_row['num_pixels'].values[0]
        baseline_total_size = effort7_total_row['baseline_size_bytes'].values[0]
        wop8_total_size = effort7_total_row['wop8_size_bytes'].values[0]
        
        # Calculate BPP for total dataset
        baseline_bpp = (baseline_total_size * 8) / total_pixels
        wop8_bpp = (wop8_total_size * 8) / total_pixels
        
        # Calculate BPP improvement and percentage based on totals
        bpp_improvement = baseline_bpp - wop8_bpp
        improvement_percentage = (effort7_total_row['size_reduction_bytes'].values[0] / baseline_total_size) * 100
        
        # Add these calculated values to the total row
        effort7_total_row['baseline_bpp'] = baseline_bpp
        effort7_total_row['wop8_bpp'] = wop8_bpp
        effort7_total_row['bpp_improvement'] = bpp_improvement
        effort7_total_row['improvement_percentage'] = improvement_percentage
        
        # Calculate totals for Effort Level 9 with correct BPP calculation
        effort9_total_row = pd.DataFrame({
            'image_name': ['TOTAL'],
            'width': [None],
            'height': [None],
            'num_pixels': [effort9_df['num_pixels'].sum()],
            'uncompressed_size_bytes': [effort9_df['uncompressed_size_bytes'].sum()],
            'baseline_size_bytes': [effort9_df['baseline_size_bytes'].sum()],
            'baseline_mae': [effort9_df['baseline_mae'].mean()],
            'wop8_size_bytes': [effort9_df['wop8_size_bytes'].sum()],
            'wop8_mae': [effort9_df['wop8_mae'].mean()],
            'size_reduction_bytes': [effort9_df['size_reduction_bytes'].sum()],
        })
        
        # Calculate BPP based on total dataset values for effort 9
        total_pixels = effort9_total_row['num_pixels'].values[0]
        baseline_total_size = effort9_total_row['baseline_size_bytes'].values[0]
        wop8_total_size = effort9_total_row['wop8_size_bytes'].values[0]
        
        # Calculate BPP for total dataset
        baseline_bpp = (baseline_total_size * 8) / total_pixels
        wop8_bpp = (wop8_total_size * 8) / total_pixels
        
        # Calculate BPP improvement and percentage based on totals
        bpp_improvement = baseline_bpp - wop8_bpp
        improvement_percentage = (effort9_total_row['size_reduction_bytes'].values[0] / baseline_total_size) * 100
        
        # Add these calculated values to the total row
        effort9_total_row['baseline_bpp'] = baseline_bpp
        effort9_total_row['wop8_bpp'] = wop8_bpp
        effort9_total_row['bpp_improvement'] = bpp_improvement
        effort9_total_row['improvement_percentage'] = improvement_percentage

        # Append total rows back to dataframes
        effort7_df = pd.concat([effort7_df, effort7_total_row], ignore_index=True)
        effort9_df = pd.concat([effort9_df, effort9_total_row], ignore_index=True)
        
        # Write updated dataframes back to Excel
        with pd.ExcelWriter(excel_path, mode='a', if_sheet_exists='replace') as writer:
            effort7_df.to_excel(writer, sheet_name='Effort Level 7', index=False)
            effort9_df.to_excel(writer, sheet_name='Effort Level 9', index=False)
        
        # Update summary sheet to include effort level results
        update_summary_with_effort_results(excel_path)
        
        return True
    except Exception as e:
        print(f"Error updating with effort level results: {e}")
        return False

def update_summary_with_effort_results(excel_path):
    """
    Update the summary sheet to include both testing vs. all images comparison
    and effort level results, with metrics calculated from totals.
    
    Args:
        excel_path (str): Path to the Excel file
        
    Returns:
        bool: True if successful, False otherwise
    """
    try:
        # Read existing Excel file
        test_df = pd.read_excel(excel_path, sheet_name='Testing')
        all_df = pd.read_excel(excel_path, sheet_name='All Images')
        effort7_df = pd.read_excel(excel_path, sheet_name='Effort Level 7')
        effort9_df = pd.read_excel(excel_path, sheet_name='Effort Level 9')
        
        # Get total rows for data extraction
        test_total = test_df[test_df['image_name'] == 'TOTAL'].iloc[0]
        all_total = all_df[all_df['image_name'] == 'TOTAL'].iloc[0]
        effort7_total = effort7_df[effort7_df['image_name'] == 'TOTAL'].iloc[0]
        effort9_total = effort9_df[effort9_df['image_name'] == 'TOTAL'].iloc[0]
        
        # Calculate BPP correctly from totals (bytes*8/pixels)
        # For Predictor Mode 6 (isolated)
        test_total_pixels = test_total['num_pixels']
        test_baseline_bpp = (test_total['baseline_size_bytes'] * 8) / test_total_pixels
        test_wop8_bpp = (test_total['wop8_size_bytes'] * 8) / test_total_pixels
        test_bpp_improvement = test_baseline_bpp - test_wop8_bpp
        test_improvement_percentage = ((test_total['baseline_size_bytes'] - test_total['wop8_size_bytes']) / test_total['baseline_size_bytes']) * 100
        
        all_total_pixels = all_total['num_pixels']
        all_baseline_bpp = (all_total['baseline_size_bytes'] * 8) / all_total_pixels
        all_wop8_bpp = (all_total['wop8_size_bytes'] * 8) / all_total_pixels
        all_bpp_improvement = all_baseline_bpp - all_wop8_bpp
        all_improvement_percentage = ((all_total['baseline_size_bytes'] - all_total['wop8_size_bytes']) / all_total['baseline_size_bytes']) * 100
        
        # For Effort Levels
        effort7_total_pixels = effort7_total['num_pixels']
        effort7_baseline_bpp = (effort7_total['baseline_size_bytes'] * 8) / effort7_total_pixels
        effort7_wop8_bpp = (effort7_total['wop8_size_bytes'] * 8) / effort7_total_pixels
        effort7_bpp_improvement = effort7_baseline_bpp - effort7_wop8_bpp
        effort7_improvement_percentage = ((effort7_total['baseline_size_bytes'] - effort7_total['wop8_size_bytes']) / effort7_total['baseline_size_bytes']) * 100
        
        effort9_total_pixels = effort9_total['num_pixels']
        effort9_baseline_bpp = (effort9_total['baseline_size_bytes'] * 8) / effort9_total_pixels
        effort9_wop8_bpp = (effort9_total['wop8_size_bytes'] * 8) / effort9_total_pixels
        effort9_bpp_improvement = effort9_baseline_bpp - effort9_wop8_bpp
        effort9_improvement_percentage = ((effort9_total['baseline_size_bytes'] - effort9_total['wop8_size_bytes']) / effort9_total['baseline_size_bytes']) * 100
        
        # Create summary dataframe with both comparisons
        summary_data = [
            ['W-OP8 Compression Results Summary', '', '', ''],
            ['', '', '', ''],
            ['Part 1: Testing vs All Images (Predictor Mode 6)', '', '', ''],
            ['Metric', 'Testing Set', 'All Images', ''],
            ['', '', '', ''],
            ['Dataset Statistics:', '', '', ''],
            ['Images Count', len(test_df) - 1, len(all_df) - 1, ''],  # Subtract 1 for total row
            ['Total Pixels', f"{test_total_pixels:,}", f"{all_total_pixels:,}", ''],
            ['', '', '', ''],
            ['Baseline Performance:', '', '', ''],
            ['Total Size (bytes)', f"{test_total['baseline_size_bytes']:,}", f"{all_total['baseline_size_bytes']:,}", ''],
            ['Bits per Pixel', f"{test_baseline_bpp:.4f}", f"{all_baseline_bpp:.4f}", ''],
            ['', '', '', ''],
            ['W-OP8 Performance:', '', '', ''],
            ['Total Size (bytes)', f"{test_total['wop8_size_bytes']:,}", f"{all_total['wop8_size_bytes']:,}", ''],
            ['Bits per Pixel', f"{test_wop8_bpp:.4f}", f"{all_wop8_bpp:.4f}", ''],
            ['', '', '', ''],
            ['Improvements:', '', '', ''],
            ['Size Reduction (bytes)', f"{test_total['size_reduction_bytes']:,}", f"{all_total['size_reduction_bytes']:,}", ''],
            ['Size Reduction (%)', f"{test_improvement_percentage:.4f}%", f"{all_improvement_percentage:.4f}%", ''],
            ['Bits per Pixel Improvement', f"{test_bpp_improvement:.4f}", f"{all_bpp_improvement:.4f}", ''],
            ['', '', '', ''],
            ['Part 2: Effort Level Comparison (All Images)', '', '', ''],
            ['Metric', 'Isolated (Predictor Mode 6)', 'Effort Level 7', 'Effort Level 9'],
            ['', '', '', ''],
            ['Dataset Statistics:', '', '', ''],
            ['Images Count', len(all_df) - 1, len(effort7_df) - 1, len(effort9_df) - 1],
            ['Total Pixels', f"{all_total_pixels:,}", f"{effort7_total_pixels:,}", f"{effort9_total_pixels:,}"],
            ['', '', '', ''],
            ['Baseline Performance:', '', '', ''],
            ['Total Size (bytes)', f"{all_total['baseline_size_bytes']:,}", f"{effort7_total['baseline_size_bytes']:,}", f"{effort9_total['baseline_size_bytes']:,}"],
            ['Bits per Pixel', f"{all_baseline_bpp:.4f}", f"{effort7_baseline_bpp:.4f}", f"{effort9_baseline_bpp:.4f}"],
            ['', '', '', ''],
            ['W-OP8 Performance:', '', '', ''],
            ['Total Size (bytes)', f"{all_total['wop8_size_bytes']:,}", f"{effort7_total['wop8_size_bytes']:,}", f"{effort9_total['wop8_size_bytes']:,}"],
            ['Bits per Pixel', f"{all_wop8_bpp:.4f}", f"{effort7_wop8_bpp:.4f}", f"{effort9_wop8_bpp:.4f}"],
            ['', '', '', ''],
            ['Improvements:', '', '', ''],
            ['Size Reduction (bytes)', f"{all_total['size_reduction_bytes']:,}", f"{effort7_total['size_reduction_bytes']:,}", f"{effort9_total['size_reduction_bytes']:,}"],
            ['Size Reduction (%)', f"{all_improvement_percentage:.4f}%", f"{effort7_improvement_percentage:.4f}%", f"{effort9_improvement_percentage:.4f}%"],
            ['Bits per Pixel Improvement', f"{all_bpp_improvement:.4f}", f"{effort7_bpp_improvement:.4f}", f"{effort9_bpp_improvement:.4f}"],
        ]
        
        # Convert to DataFrame
        summary_df = pd.DataFrame(summary_data)
        
        # Write to Excel
        with pd.ExcelWriter(excel_path, mode='a', if_sheet_exists='replace') as writer:
            # Write summary sheet
            summary_df.to_excel(writer, sheet_name='Summary', index=False, header=False)
            
            # Access the workbook and set column widths
            worksheet = writer.sheets['Summary']
            worksheet.column_dimensions['A'].width = 30
            worksheet.column_dimensions['B'].width = 25
            worksheet.column_dimensions['C'].width = 25
            worksheet.column_dimensions['D'].width = 25
            
            # Style header row and section titles
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Title styling
            worksheet.cell(row=1, column=1).font = Font(bold=True, size=14)
            worksheet.merge_cells('A1:D1')
            worksheet.cell(row=1, column=1).alignment = Alignment(horizontal='center')
            
            # Part 1 title
            worksheet.cell(row=3, column=1).font = Font(bold=True, size=12)
            worksheet.merge_cells('A3:D3')
            
            # Part 1 header row
            for col in range(1, 4):
                cell = worksheet.cell(row=4, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # Part 2 title
            worksheet.cell(row=23, column=1).font = Font(bold=True, size=12)
            worksheet.merge_cells('A23:D23')
            
            # Part 2 header row (this is the important fix)
            for col in range(1, 5):
                cell = worksheet.cell(row=24, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # Style section headers for both parts
            section_rows_part1 = [6, 10, 14, 18]
            section_rows_part2 = [27, 31, 35, 39]
            
            for row in section_rows_part1 + section_rows_part2:
                worksheet.cell(row=row, column=1).font = Font(bold=True)
            
  
            
            for row in range(1, 43):  # Adjust based on actual content
                for col in range(1, 5):
                    cell = worksheet.cell(row=row, column=col)
                    # cell.border = thin_border
        
        return True
    except Exception as e:
        print(f"Error updating summary with effort results: {e}")
        return False